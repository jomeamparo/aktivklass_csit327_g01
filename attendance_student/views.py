import json
from collections import defaultdict

import openpyxl
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, render
from django.utils.encoding import smart_str
from django.views.decorators.http import require_GET
from openpyxl.utils import get_column_letter

from core.models import (ActivityRecord, Attendance, Class, QuizGrade,
                         Student)

# ==================================
# SHARED/GENERAL VIEWS
# ==================================

def student_dashboard(request):
    student_id = request.session.get('user_id')
    role = request.session.get('role')
    if role != 'student' or not student_id:
        return render(request, 'error.html', {'message': 'You are not logged in as a student.'})
    try:
        student = Student.objects.get(student_id=student_id)
    except Student.DoesNotExist:
        return render(request, 'error.html', {'message': 'No student record found for this account.'})

    # Attendance Rate
    total_attendance = Attendance.objects.filter(student=student).count()
    present_attendance = Attendance.objects.filter(student=student, status='Present').count()
    attendance_rate = round((present_attendance / total_attendance) * 100, 1) if total_attendance > 0 else 0

    # Average Grade (from QuizGrade and ActivityRecord)
    quiz_grades = QuizGrade.objects.filter(student=student)
    activity_records = ActivityRecord.objects.filter(student=student)
    quiz_scores = list(quiz_grades.values_list('percentage', flat=True))
    activity_scores = [((rec.score / rec.perfect_score) * 100) for rec in activity_records if rec.score is not None and rec.perfect_score and rec.perfect_score > 0]
    all_scores = quiz_scores + activity_scores
    average_grade = round(sum(all_scores) / len(all_scores), 1) if all_scores else None

    # Most Recent Grade (QuizGrade)
    recent_grade = quiz_grades.order_by('-graded_at').first()

    # Recent Activities (ActivityRecord, last 5)
    recent_activities = activity_records.order_by('-date')[:5]

    context = {
        'attendance_rate': attendance_rate,
        'average_grade': average_grade,
        'recent_grade': recent_grade,
        'recent_activities': recent_activities,
        'role': role,
        'fullname': f"{student.first_name} {student.last_name}",
        'email': student.email,
        'avatar_url': student.profile.avatar.url if hasattr(student, 'profile') and student.profile.avatar else None,
    }
    return render(request, 'attendance_student/student_dashboard.html', context)


# ==================================
# ATTENDANCE VIEWS
# ==================================

def student_attendance_dashboard(request):
    student_id = request.session.get('user_id')
    role = request.session.get('role')
    if role != 'student' or not student_id:
        return render(request, 'error.html', {'message': 'You are not logged in as a student.'})
    try:
        student = Student.objects.get(student_id=student_id)
    except Student.DoesNotExist:
        return render(request, 'error.html', {'message': 'No student record found for this account.'})

    attendance_records = Attendance.objects.filter(student=student).select_related('class_obj')
    total_attendance = attendance_records.count()

    class_cards = {}
    stats_dict = {}
    for record in attendance_records:
        class_obj = record.class_obj
        if class_obj.id not in class_cards:
            class_cards[class_obj.id] = {
                'id': class_obj.id,
                'name': class_obj.subject_name,
                'code': class_obj.class_code,
                'count': 0
            }
        class_cards[class_obj.id]['count'] += 1

        class_label = str(record.class_obj)
        if class_label not in stats_dict:
            stats_dict[class_label] = {'subject': class_label, 'present': 0, 'absent': 0}
        if record.status == 'Present':
            stats_dict[class_label]['present'] += 1
        elif record.status == 'Absent':
            stats_dict[class_label]['absent'] += 1

    context = {
        'total_attendance': total_attendance,
        'class_cards': list(class_cards.values()),
        'attendance_stats_json': json.dumps(list(stats_dict.values())),
        'role': role,
        'fullname': f"{student.first_name} {student.last_name}",
        'email': student.email,
        'avatar_url': student.profile.avatar.url if hasattr(student, 'profile') and student.profile.avatar else None,
    }
    return render(request, 'attendance_student/student_attendance_dashboard.html', context)

@require_GET
def student_attendance_subject_detail(request, class_id):
    student_id = request.session.get('user_id')
    role = request.session.get('role')
    if role != 'student' or not student_id:
        return render(request, 'error.html', {'message': 'Invalid request.'})
    try:
        student = Student.objects.get(student_id=student_id)
        class_obj = get_object_or_404(Class, id=class_id)
    except Student.DoesNotExist:
        return render(request, 'error.html', {'message': 'No student record found for this account.'})

    attendance_records = Attendance.objects.filter(student=student, class_obj=class_obj)
    present_count = attendance_records.filter(status='Present').count()
    absent_count = attendance_records.filter(status='Absent').count()

    pie_chart_data = {'present': present_count, 'absent': absent_count}

    context = {
        'class_obj': class_obj,
        'attendance_records': attendance_records,
        'pie_chart_data_json': json.dumps(pie_chart_data),
        'role': role,
        'fullname': f"{student.first_name} {student.last_name}",
        'email': student.email,
        'avatar_url': student.profile.avatar.url if hasattr(student, 'profile') and student.profile.avatar else None,
    }
    return render(request, 'attendance_student/attendance_subject_detail.html', context)


def student_attendance_all(request):
    student_id = request.session.get('user_id')
    role = request.session.get('role')
    if role != 'student' or not student_id:
        return render(request, 'error.html', {'message': 'You are not logged in as a student.'})
    try:
        student = Student.objects.get(student_id=student_id)
    except Student.DoesNotExist:
        return render(request, 'error.html', {'message': 'No student record found for this account.'})

    context = {
        'attendance_records': Attendance.objects.filter(student=student).select_related('class_obj').order_by('-date'),
        'role': role,
        'fullname': f"{student.first_name} {student.last_name}",
        'email': student.email,
        'avatar_url': student.profile.avatar.url if hasattr(student, 'profile') and student.profile.avatar else None,
    }
    return render(request, 'attendance_student/attendance_detail.html', context)



# ==================================
# RESULTS/GRADES VIEWS
# ==================================

def student_results_dashboard(request):
    student_id = request.session.get('user_id')
    role = request.session.get('role')
    if role != 'student' or not student_id:
        return render(request, 'error.html', {'message': 'Invalid request.'})
    try:
        student = Student.objects.get(student_id=student_id)
    except Student.DoesNotExist:
        return render(request, 'error.html', {'message': 'No student record found.'})

    subject_data = defaultdict(lambda: {'scores': [], 'count': 0, 'class_id': None})

    # 1. Aggregate Quiz Grades
    quiz_grades = QuizGrade.objects.filter(student=student).select_related('quiz__class_obj')
    for grade in quiz_grades:
        if grade.quiz and grade.quiz.class_obj:
            class_obj = grade.quiz.class_obj
            subject_name = class_obj.subject_name
            subject_data[subject_name]['scores'].append(grade.percentage)
            subject_data[subject_name]['count'] += 1
            subject_data[subject_name]['class_id'] = class_obj.id

    # 2. Aggregate Activity Records
    activity_records = ActivityRecord.objects.filter(student=student).select_related('enrollment__enrolled_class')
    for record in activity_records:
        if record.score is not None and record.perfect_score and record.perfect_score > 0:
            class_obj = record.enrollment.enrolled_class
            subject_name = class_obj.subject_name
            percentage = (record.score / record.perfect_score) * 100
            subject_data[subject_name]['scores'].append(percentage)
            subject_data[subject_name]['count'] += 1
            subject_data[subject_name]['class_id'] = class_obj.id

    # 3. Prepare data for template
    subject_cards = []
    chart_data = []
    for name, data in subject_data.items():
        avg_score = round(sum(data['scores']) / len(data['scores']), 1) if data['scores'] else 0
        subject_cards.append({
            'subject_name': name,
            'average_score': avg_score,
            'total_activities': data['count'],
            'class_id': data['class_id']
        })
        chart_data.append({'subject': name, 'average': avg_score})

    context = {
        'subject_cards': subject_cards,
        'chart_data_json': json.dumps(chart_data),
        'role': role,
        'fullname': f"{student.first_name} {student.last_name}",
        'email': student.email,
        'avatar_url': student.profile.avatar.url if hasattr(student, 'profile') and student.profile.avatar else None,
    }
    return render(request, 'attendance_student/student_results_dashboard.html', context)


def results_subject_detail(request, class_id):
    student_id = request.session.get('user_id')
    role = request.session.get('role')
    if role != 'student' or not student_id:
        return render(request, 'error.html', {'message': 'Invalid request.'})
    try:
        student = Student.objects.get(student_id=student_id)
        class_obj = get_object_or_404(Class, id=class_id)
    except Student.DoesNotExist:
        return render(request, 'error.html', {'message': 'No student record found.'})

    # Note: Adjust the quiz grade filter based on your actual model relationships
    # This assumes a path like: QuizGrade -> Quiz -> Class
    activity_records = ActivityRecord.objects.filter(student=student, enrollment__enrolled_class=class_obj)
    quiz_grades = QuizGrade.objects.filter(student=student, quiz__class_obj=class_obj)

    context = {
        'class_obj': class_obj,
        'activity_records': activity_records,
        'quiz_grades': quiz_grades,
        'role': role,
        'fullname': f"{student.first_name} {student.last_name}",
        'email': student.email,
        'avatar_url': student.profile.avatar.url if hasattr(student, 'profile') and student.profile.avatar else None,
    }
    return render(request, 'attendance_student/results_subject_detail.html', context)

# ==================================
# EXPORT VIEWS
# ==================================

def export_attendance_excel(request):
    student_id = request.session.get('user_id')
    # ... (Add student check)
    student = Student.objects.get(student_id=student_id)
    attendance_records = Attendance.objects.filter(student=student).select_related('class_obj')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Attendance Records'
    headers = ['Date', 'Class', 'Status', 'Feedback']
    ws.append(headers)

    for record in attendance_records:
        ws.append([
            record.date.strftime('%Y-%m-%d'),
            str(record.class_obj),
            record.status,
            record.feedback or ''
        ])

    for i, col in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = 20

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"attendance_{student.student_id}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{smart_str(filename)}"'
    wb.save(response)
    return response